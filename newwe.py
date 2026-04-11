from openpyxl import Workbook
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

#--- Custom exceptions ---

class InvalidRowError(Exception):
    """Triggered when a row of data could not be converted into the right format"""

class MissingDataError(Exception): 
    """Triggered when the necessary data set was not available before processing"""

# --- Private Helper/ Data Parsing ---

def _parse_row(row: list) -> dict | None: 
    """
    Parses a raw data into a structured dict
    
    Expected format: [name: str, date: 'YYYY-MM-DD', grade1, grade2, ...]
    
    Raises: InvalidRowError: if the row is distorted or contains bad values
    """
    if not len(row) or len(row) < 3: 
        raise InvalidRowError(f"Row too short to contain name, data and at least one grade: {row!r}")
    
    try: 
        name = row[0]
        date = datetime.strptime(row[1], '%Y-%m-%d').date()
        grades = [g for g in row[2:] if isinstance(g, (int, float))]
    except ValueError as e: 
        raise InvalidRowError(f"Bad date format in row {row!r}")

    if not grades: 
        raise InvalidRowError(f"No numeric grades found in row {row!r}")

    return {"name": name, "date": date, "grades": grades}

# --- Main Class / Builder ---

class ExcelReportBuilder: 
    """
    Builds a multi-sheet Excel workbook from a list of student records.

    Usage::

        wb = ExcelReportBuilder()
        wb.add_workbook(students, headers=["Name", "Date", "grade1", "grade2"])
        wb.create_summary(fail_under=70, file="report.xlsx")
    """   
    def __init__(self) -> None: 
        self._workbook = Workbook()
        #Remove current active sheet so sheets will be added in our terms
        self._workbook.remove(self._workbook.active)
        self._parsed_records = []
    
    # ----- Public API -----

    def add_workbook(self, data: list[list], headers: list[str] | None = None) -> None: 
        """
        Parse 'data' with the helper func and write it to a "Main Report" sheet

        Args:
            data:    Each inner list must be [name, 'YYYY-MM-DD', grade, ...]
            headers: Optional column headers. Auto-generated when not given

        Raises:
            MissingDataError: if 'data' is empty.
            InvalidRowError:  Originates from func: '_parse_row()' for bad rows
        """
        sheet = self._workbook.create_sheet("Main Report")

        if not data: 
            raise MissingDataError("Data must be a non-empty list")
        
        if headers: 
            sheet.append(headers)
        else: 
            width = max(len(row) for row in data)
            sheet.append(["Names", "Dates"] + [f"Grade {i}" for i in range(1, width - 1)])

        count: int = 0
        for row in data: 
            parsed = _parse_row(row) #Raises InvalidRowError on bad input
            sheet.append([parsed["name"], parsed['date'], *parsed['grades']])
            count += 1
            self._parsed_records.append(parsed) #So that we don't need to waste time and energy on parsing again
        logger.info("Records sheet: Wrote %d rows.", count)
       
    
    def create_summary(self, fail_under: int = 70, file: str = "student_report.xlsx") -> None: 
        """
        Calculate each student's average, add a “Summary” sheet, and save the file.

        Args:
            fail_under: Average below this value is marked "FAIL"
            file:       Output file path (must end in .xlsx) which would be called from following private func

        Raises:
            MissingDataError: if func: 'add_workbook()' has not been called yet
            InvalidRowError:  Originates from func: '_parse_row()' for bad rows
            PermissionError:  if the file is open in another process
        """
        if not hasattr(self, "_parsed_records"): 
            raise MissingDataError("Call add_workbook() before create_summary().")

        sheet = self._workbook.create_sheet("Summary")
        sheet.append(["Names", "Dates", "Average Grade", "Status"])

        count: int = 0
        for parsed in self._parsed_records: 
            # 'parsed' is already a dict, so we jump straight to the math
            average_grade = round(sum(parsed["grades"]) / len(parsed["grades"]), 1)
            status: str = "PASS" if average_grade >= fail_under else "FAIL"
            
            sheet.append([parsed['name'], parsed['date'], average_grade, status])
            count += 1
        
        logger.info("Summary sheet: summarised %d students.", count)
        self._save(file)
    
    # --- Private Helper ---
    def _save(self, file: str) -> None: 
        if not file.endswith('.xlsx'): 
            file += '.xlsx'
        try:
            self._workbook.save(file)
            logger.info("Workbook saved to %s.", file)
        except PermissionError: 
            logger.critical("Couldn't save %s. Please close the file and try again", file)
            raise

if __name__ == "__main__": 
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    
    #example list of student names, you can change it if you want to try out the error handling :)
    students = [
        ["Jane", "2024-06-20", 40, 70, 100, 56],
        ["Bob",  "2026-06-20", 93, 56,  74, 89]
    ]

    wb = ExcelReportBuilder() #Initiate the object
    wb.add_workbook(students) #First func for "Report" sheet
    wb.create_summary(fail_under=60, file="Student_Summaries.xlsx") #final function for "Summary" sheet



